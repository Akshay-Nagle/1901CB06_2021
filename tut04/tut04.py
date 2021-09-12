#The program might take some time to create other .xlsx files if initial csv file is big. Thank you for your patience.



import os                                                                                           #importing modules
import csv
import openpyxl

#function to convert .csv file to .xlsx file
def convert():
    wb = openpyxl.Workbook()                                                                        #calling Workbook() function to create a new blank workbook
    Sheet = wb.active                                                                               #getting workbook active

    f = open("F:/5th sem/Python/1901CB06_2021/tut04/regtable_old.csv","r")                          #giving the location of the file to open in read mode
    data = csv.reader(f)                                                                            #reader() of csv used to read all details and storing into data
    for rows in data:                                                                               #loop toappend details into sheet
        Sheet.append(rows)
    Sheet.title = "Main"                                                                            #renaming the title of sheet
    wb.save('F:/5th sem/Python/1901CB06_2021/tut04/regtable_old.xlsx')                              #saving changes made

#function to generate all the required files
def general(index):         
    convert()                                                                                       #calling finction convert()
    header = ["rollno", "register_sem" ,"sub_no","sub_type"]                                        #initializing header of .xlsx files
    wb = openpyxl.load_workbook("F:/5th sem/Python/1901CB06_2021/tut04/regtable_old.xlsx")          #to open the workbook, wb is created
    Sheet = wb.active                                                                               #getting workbook active
    data = list(Sheet.iter_rows(values_only=True))                                                  #to store all of the rows in the file as lists of row values
    cell_data = ()                                                                                  #intializing variable

    for m in data:                                                                                  #loop to select details from list data
        cell_data = (m[0], m[1], m[3], m[8])                                                        #storing desired values
        if index == 0:                                                                              #condition to generate required files with the help of index
            dir = "output_individual_roll/"
            num = str(m[0])                                                                         
        else:
            dir = "output_by_subject/"
            num = str(m[3])    
                                                             
        route = "F:/5th sem/Python/1901CB06_2021/tut04/" + dir + num + ".xlsx"                      #path in which file is going to be store
        
        if os.path.isfile(route) == False:                                                          #checking whether that file is  already present or not
            wb = openpyxl.Workbook()
            Sheet = wb.active
            Sheet.append(header)                                                                    #appending intial values to .xlsx file
            Sheet.title = num                                                                       #renaming the title of sheet                                     
            wb.save(route)
    
        wb = openpyxl.load_workbook(route)                                                          
        Sheet = wb.active
        Sheet.title = num
        Sheet.append(cell_data)                                                                     #appending required values in file
        wb.save(route)
     


#function to generate file by subject
def output_by_subject():                
    general(3)                                                                                      #passing arguement "3"
    return


#function to generate file by individual rollnos
def output_individual_roll():
    general(0)                                                                                      #passing arguement "0"
    return


#driver's code
output_by_subject()
output_individual_roll()
