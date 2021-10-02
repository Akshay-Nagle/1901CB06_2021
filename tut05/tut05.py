import os
import csv
import openpyxl

#opening csv files
Grad = open("F:/5th sem/Python/1901CB06_2021/tut05/grades.csv")                                             
Name_Roll = open("F:/5th sem/Python/1901CB06_2021/tut05/names-roll.csv")
Subj_mas = open("F:/5th sem/Python/1901CB06_2021/tut05/subjects_master.csv")

gDIC = {"AA": 10, "AB": 9, "BB": 8, "BC": 7, "CC": 6, "CD": 5, "DD": 4, "F": 0, "I": 0}                     #grade dictionary
header = ["S No", "Sub Code", "Sub Name", "LTP", "Cred", "Type", "Grade"]                                   #row header

#function to calculate spi and cpi
def OverallResult(rollNum: str):
    name = ["Name of Student"]                                                                              #intialisation of variables
    spi, cpi = ["SPI"], ["CPI"]
    semCreds, TCreds  = ["Semester wise Credit Taken"], ["Total Credits Taken"]
    sRow = ["Semester No", 1]

    mSEM = 0
    for ind, line in enumerate(Grad.readlines()):                                                           #loop condition for calculating maximum semester for given rollno
        if line[0] == rollNum:
            mSEM = max(mSEM, int(line[1]))

    for ind, line in enumerate(Name_Roll.readlines()):                                                      #loop condition for appending individual names in name list
        if ind > 0 and line[0] == rollNum:
            name.append(line[1])

    mSEM = mSEM + 1

    for f in range(1, mSEM):                                                                                #loop condition for calculating credits and Tgrade
        credits, Tgrade = 0, 0
        for index, line in enumerate((csv.reader(Grad))):
            if index > 0 and line[0] == rollNum:  
                    if (int(line[1]) == f):
                        credits = credits + int(line[3])
                        finalGrade = asteriskError(line[4].strip()) 
                        Tgrade = Tgrade + (int(line[3]) * gDIC[finalGrade])
        
        if credits == 0:                                                                                    #if condition to check value of credits
            semCreds.append(0)
            spi.append(0)
        else:
            spi.append((Tgrade/credits).__round__(2))                                                       #using round function to get required SPI 
            semCreds.append(credits)

    tillCPI = spi[1] * semCreds[1]                                                                          #Calculating 1st CPI
    tillCreds = semCreds[1]
    TCreds.append(tillCreds)
    cpi.append(spi[1])

    for k in range(2, mSEM):                                                                                #loop condition to calculate CPI
        sRow.append(k)
        tillCreds = tillCreds + semCreds[k]                                                                 #calculating credits
        TCreds.append(tillCreds)
        tillCPI = tillCPI + (spi[k] * semCreds[k])                                                          #calculating till now CPI
        cpi.append((tillCPI / tillCreds).__round__(2))                                                      #calculating overall CPI by using round function

    return name, sRow, semCreds, TCreds, spi, cpi                                                           #returning required values


#function to remove asterisk from grade end
def asteriskError(grade):
    if grade[:-1] == '*' :
        grade.replace('*', '')
    return grade


#Function to create required sheets and displaying results
def prepSemSheets():
    for index, line in enumerate(Grad.readlines()):                                                         #loop condition to stored reuired data in data variable
        data = [1, line[2], "", "", line[3], line[5], line[4].strip()]
        
        route = "F:/5th sem/Python/1901CB06_2021/tut05/output/" + str(line[0]) + ".xlsx"                                      #defining path to store the output files or results

        for index, line in enumerate(csv.reader(Subj_mas)):                                                 #loop condition to check rollno and storing misssing required data
            if line[0] == data[1]:
                data[2] = line[1]
                data[3] = line[2]

        if index > 0:                                           
            if os.path.exists(route) == False:                                                              #checking whether file already existed or not
                wb = openpyxl.Workbook()
                wb.save(route)                                                                              #if not present then creating and then store one file

            reqSheet = "Sem " + str(line[1])                                                                #saving sheet name for later purposes
            existing_workbook = openpyxl.load_workbook(route)                                               #accessing previously created sheet

            if "Overall" in existing_workbook.sheetnames == False:                                          #cheking whether Overall named sheet present or not                                                      
                del existing_workbook["Sheet"]                                                              #if sheet is not present then it will create 1 overall named sheet
                reqSheetO = "Overall"
                existing_workbook.create_sheet(reqSheetO)
                sheet = existing_workbook[reqSheetO]
                nameRow, sRow, semCreds, TCreds, spi, cpi = OverallResult(line[0])
                sheet.append(["Roll No", line[0]])                                                          #appending required values in the sheet
                sheet.append(nameRow)
                sheet.append(["Discipline", str(str(line[0])[4] + str(line[0])[5])])
                sheet.append(sRow)
                sheet.append(semCreds)
                sheet.append(spi)
                sheet.append(TCreds)
                sheet.append(cpi)
                existing_workbook.save(route)

            if reqSheet in existing_workbook.sheetnames == False:                                           #if sheet is not present then it will create one
                existing_workbook.create_sheet(reqSheet)
                existing_workbook[reqSheet].append(header)                                                  #it will append header
                existing_workbook.save(route)                                                               #saving the sheet

            active_workbook = existing_workbook
            data[0] = active_workbook[reqSheet].max_row                                                     #calculating max rows
            active_workbook[reqSheet].append(data)                                                          #adding / appending data in respective sheets
            existing_workbook.save(route)                                                                   #saving the sheet
    
    return                                                                                                         

#Driver's code
prepSemSheets()
